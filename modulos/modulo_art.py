import tkinter as tk
import tkinter.ttk as ttk
from tkinter import messagebox, filedialog
import customtkinter as ctk
import threading
import logging
import mysql.connector
from mysql.connector import Error
from datetime import datetime, timedelta
from PIL import Image, ImageTk, ImageDraw
import time
import json
from concurrent.futures import ThreadPoolExecutor
import traceback
from dotenv import load_dotenv
from io import BytesIO
from tkcalendar import DateEntry
import os
import sys

# Agregar el directorio raíz al path para poder importar los módulos
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

# Importaciones de utils
from utils.thread_manager import ThreadManager, DatabasePool
from utils.interface_manager import EstiloApp

# Cargar variables de entorno
load_dotenv()

# Configuración de tema y colores
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

class EstiloApp:
    """Clase para definir los colores y estilos de la aplicación"""
    COLOR_PRINCIPAL = "#E3F2FD"
    COLOR_SECUNDARIO = "#90CAF9"
    COLOR_TEXTO = "#000000"
    COLOR_FRAMES = "#BBDEFB"
    COLOR_HEADER = "#FFFFFF"

    # Nuevos colores para botones CRUD
    BOTON_INSERTAR = "#4CAF50"  # Verde - representa creación/nuevo
    BOTON_INSERTAR_HOVER = "#45A049"  # Verde más oscuro
    
    BOTON_MODIFICAR = "#2196F3"  # Azul - representa actualización
    BOTON_MODIFICAR_HOVER = "#1976D2"  # Azul más oscuro
    
    BOTON_ELIMINAR = "#F44336"  # Rojo - representa eliminación/peligro
    BOTON_ELIMINAR_HOVER = "#D32F2F"  # Rojo más oscuro
    
    BOTON_LIMPIAR = "#757575"  # Gris - representa acción neutral
    BOTON_LIMPIAR_HOVER = "#616161"  # Gris más oscuro

class DialogManager:
    """Clase para gestionar diálogos y ventanas modales"""
    def __init__(self):
        self.active_dialogs = []
    
    def show_dialog(self, dialog_class, *args, **kwargs):
        """Muestra un diálogo modal"""
        dialog = dialog_class(*args, **kwargs)
        self.active_dialogs.append(dialog)
        return dialog
    
    def close_all(self):
        """Cierra todos los diálogos activos"""
        for dialog in self.active_dialogs:
            try:
                dialog.destroy()
            except:
                pass
        self.active_dialogs = []

class AplicacionART:
    def __init__(self, parent_frame=None):
        """
        Inicializar la aplicación
        :param parent_frame: Frame padre donde se mostrará el módulo
        """
        self.parent = parent_frame
        self.is_standalone = parent_frame is None
        
        # Primero inicializar el root correctamente
        if self.is_standalone:
            self.root = ctk.CTk()
            self.main_container = self.root
        else:
            # Si tenemos un parent_frame, usarlo como contenedor
            self.root = self._find_root_window(parent_frame)
            self.main_container = ctk.CTkFrame(parent_frame, fg_color=EstiloApp.COLOR_PRINCIPAL)
            self.main_container.grid(row=0, column=0, sticky="nsew")
        
        # Configurar el sistema de logging primero
        self.logger = self._setup_logging()
        
        # Inicializar otros componentes después de tener un root válido
        self.dialog_manager = DialogManager()
        self.db_pool = DatabasePool()
        self.is_destroyed = False
        self.accidente_seleccionado_id = None
        self.actualizando_treeview = False
        self.lock = threading.Lock()
        
        # Iniciar la interfaz de forma asíncrona
        if self.is_standalone:
            self.root.after(100, self._init_async)
        else:
            self._init_async()
        
    def _setup_logging(self):
        """Configurar el sistema de logging"""
        logger = logging.getLogger('AplicacionART')
        logger.setLevel(logging.DEBUG)
        
        # Crear manejador para la consola
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.DEBUG)
        
        # Crear formato para los logs
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        console_handler.setFormatter(formatter)
        
        # Agregar manejador al logger
        logger.addHandler(console_handler)
        
        # Crear directorio de logs si no existe
        os.makedirs('logs', exist_ok=True)
        
        # Crear manejador para archivo
        file_handler = logging.FileHandler('logs/aplicacion_art.log')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)
        
        # Agregar manejador al logger
        logger.addHandler(file_handler)
        
        return logger
    
    def setup_window(self):
        """Configuración inicial de la ventana"""
        if self.is_standalone:
            self.root.title("Sistema de Gestión de ART")
            self.root.state('zoomed')  # Maximizar la ventana en Windows
        
        # Configurar el frame principal independientemente del modo
        if isinstance(self.root, ctk.CTk):
            self.root.configure(fg_color=EstiloApp.COLOR_PRINCIPAL)
        else:
            # Si es un frame, solo configuramos el color
            self.root.configure(fg_color=EstiloApp.COLOR_PRINCIPAL)
            
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(0, weight=1)
    
    def create_gui(self):
        """Crear la interfaz gráfica usando grid consistentemente"""
        # Configurar el main_container primero
        self.main_container.grid_columnconfigure(0, weight=1)
        self.main_container.grid_rowconfigure(0, weight=0)  # Header
        self.main_container.grid_rowconfigure(1, weight=0)  # Separator (opcional)
        self.main_container.grid_rowconfigure(2, weight=1)  # Content
        
        # Header frame (row 0)
        header_frame = ctk.CTkFrame(
            self.main_container,
            fg_color=EstiloApp.COLOR_HEADER,
            corner_radius=10,
            height=150  # Altura fija
        )
        header_frame.grid(row=0, column=0, sticky="new", padx=35, pady=(5, 0))
        header_frame.pack_propagate(False)
        
        # Logo frame
        logo_frame = ctk.CTkFrame(
            header_frame, 
            fg_color='transparent',
            width=150,
            height=150
        )
        logo_frame.grid(row=0, column=0, rowspan=3, padx=(10, 20), pady=0)
        logo_frame.grid_propagate(False)
        
        # Logo label
        self.logo_label = ctk.CTkLabel(
            logo_frame,
            text="",
            width=150,
            height=150
        )
        self.logo_label.place(relx=0.5, rely=0.5, anchor="center")
        
        # Cargar el logo
        self._load_gif_frames()
        
        # Títulos
        titles_frame = ctk.CTkFrame(header_frame, fg_color='transparent')
        titles_frame.grid(row=0, column=1, rowspan=3, sticky="nsw", padx=(0, 20))
        
        ctk.CTkLabel(
            titles_frame,
            text="Módulo de ART - RRHH",
            font=('Roboto', 24, 'bold'),
            text_color=EstiloApp.COLOR_TEXTO
        ).grid(row=0, column=0, sticky="w", pady=(30, 5))
        
        ctk.CTkLabel(
            titles_frame,
            text="Entheus Seguridad",
            font=('Roboto', 16, 'bold'),
            text_color=EstiloApp.COLOR_SECUNDARIO
        ).grid(row=1, column=0, sticky="w", pady=5)
        
        ctk.CTkLabel(
            titles_frame,
            text="© 2025 Todos los derechos reservados",
            font=('Roboto', 12),
            text_color=EstiloApp.COLOR_TEXTO
        ).grid(row=2, column=0, sticky="w", pady=(5, 30))
        
        # Content frame
        self.content_frame = ctk.CTkFrame(
            self.main_container,
            fg_color=EstiloApp.COLOR_PRINCIPAL
        )
        self.content_frame.grid(row=2, column=0, sticky="nsew", padx=20, pady=(0, 15))
        
        # Configurar content_frame
        self.content_frame.grid_columnconfigure(0, weight=1)
        self.content_frame.grid_rowconfigure(0, weight=0)  # Form
        self.content_frame.grid_rowconfigure(1, weight=1)  # Table
        
        # Asignar main_frame
        self.main_frame = self.content_frame
        
        # Crear componentes
        self._create_form()
        self._create_table()
        
        return self.main_frame
    
    def _create_form(self):
        """Crear formulario de ART"""
        form_frame = ctk.CTkFrame(
            self.main_frame, 
            fg_color=EstiloApp.COLOR_FRAMES,
            corner_radius=10
        )
        form_frame.grid(row=0, column=0, sticky="ew", padx=15, pady=(5, 5))
        form_frame.grid_columnconfigure(1, weight=1)
        form_frame.grid_columnconfigure(2, weight=0)

        # Título del formulario
        title_label = ctk.CTkLabel(
            form_frame,
            text="Registre o Modifique Accidentes de Trabajo",
            font=('Roboto', 20, 'bold')
        )
        title_label.grid(row=0, column=0, columnspan=2, pady=(20, 20))

        # Crear campos del formulario
        self.crear_campos_formulario(form_frame)

        # Panel central con foto y datos del empleado
        self._create_employee_info_frame(form_frame)

        # Panel derecho (botones CRUD)
        self._create_crud_buttons(form_frame)
    
    def _create_employee_info_frame(self, parent):
        """Crear panel central con foto y datos del empleado"""
        info_frame = ctk.CTkFrame(parent, fg_color="transparent")
        info_frame.grid(row=1, column=1, rowspan=5, padx=20, pady=5, sticky="n")

        # Frame para la foto
        photo_frame = ctk.CTkFrame(
            info_frame,
            fg_color=EstiloApp.COLOR_PRINCIPAL,
            corner_radius=10,
            border_width=2,
            border_color=EstiloApp.COLOR_SECUNDARIO,
            width=200,
            height=220
        )
        photo_frame.pack(pady=(0, 10))
        photo_frame.pack_propagate(False)

        # Canvas para la foto
        self.photo_canvas = ctk.CTkCanvas(
            photo_frame,
            width=200,
            height=200,
            bg=EstiloApp.COLOR_PRINCIPAL,
            highlightthickness=0
        )
        self.photo_canvas.place(relx=0.5, rely=0.5, anchor="center")

        # Placeholder inicial
        self.photo_canvas.create_oval(
            50, 50, 150, 150,
            fill="#E0E0E0",
            outline=""
        )
        self.photo_canvas.create_text(
            100, 100,
            text="👤",
            font=("Arial", 40),
            fill="#909090"
        )
        self.photo_canvas.create_text(
            100, 170,
            text="Sin foto",
            font=("Arial", 10),
            fill="#606060"
        )

        # Frame para datos del empleado (debajo de la foto)
        data_frame = ctk.CTkFrame(info_frame, fg_color="transparent")
        data_frame.pack(pady=(5, 0))

        # Labels de información
        self.nombre_completo_label = ctk.CTkLabel(
            data_frame,
            text="Apellido y Nombre: -",
            font=ctk.CTkFont(size=14, weight="bold"),
            anchor="center",
            width=250
        )
        self.nombre_completo_label.pack(pady=(0, 5))

        self.total_accidentes_label = ctk.CTkLabel(
            data_frame,
            text="Total Accidentes: 0",
            font=ctk.CTkFont(size=14),
            anchor="center",
            width=250
        )
        self.total_accidentes_label.pack()
        
        # Agregar labels estadísticos adicionales
        self.ultimo_accidente_label = ctk.CTkLabel(
            data_frame,
            text="⏱️ Último accidente: No registrado",
            font=ctk.CTkFont(size=14),
            anchor="center",
            width=250
        )
        self.ultimo_accidente_label.pack(pady=(5, 0))
        
        self.dias_totales_label = ctk.CTkLabel(
            data_frame,
            text="📊 Total días con ART: 0 días",
            font=ctk.CTkFont(size=14),
            anchor="center",
            width=250
        )
        self.dias_totales_label.pack(pady=(5, 0))
    
    def crear_campos_formulario(self, parent):
        """Crear campos del formulario"""
        # Frame contenedor del formulario
        form_container = ctk.CTkFrame(parent, fg_color="transparent")
        form_container.grid(row=1, column=0, padx=20, pady=5, sticky="nsew")
        
        form_container.grid_columnconfigure(1, weight=1)
        form_container.grid_columnconfigure(3, weight=1)
        form_container.grid_columnconfigure(5, weight=0)

        # Definir campos y sus posiciones
        fields = [
            ("Legajo:", "entry_legajo", 0, 0, None),
            ("Fecha Accidente:", "entry_fecha_accidente", 1, 0, "date"),
            ("Fecha Alta:", "entry_fecha_alta", 2, 0, "date"),
            ("Diagnóstico:", "entry_diagnostico", 0, 2, None, 3),
            ("Ámbito:", "entry_ambito", 1, 2),
            ("N° Siniestro:", "entry_n_siniestro", 2, 2),
            ("Descripción:", "text_descripcion", 3, 0, "text", 2)
        ]

        # Estilo para DateEntry
        style = ttk.Style()
        style.configure(
            'Custom.DateEntry',
            fieldbackground=EstiloApp.COLOR_PRINCIPAL,
            background=EstiloApp.COLOR_SECUNDARIO,
            foreground='black',
            arrowcolor='black',
            font=('Roboto', 14)
        )

        # Crear campos
        for field in fields:
            # Label
            label = ctk.CTkLabel(
                form_container,
                text=field[0],
                font=('Roboto', 14, 'bold')
            )
            label.grid(row=field[2], column=field[3], padx=(20, 10), pady=5, sticky="e")

            # Widget
            if len(field) > 4 and field[4] == "date":
                # Frame contenedor para el DateEntry
                date_frame = ctk.CTkFrame(
                    form_container,
                    fg_color=EstiloApp.COLOR_PRINCIPAL,
                    height=35,
                    width=200,
                    border_width=2,
                    border_color=EstiloApp.COLOR_SECUNDARIO,
                    corner_radius=8
                )
                date_frame.grid(row=field[2], column=field[3]+1, sticky="w", padx=20, pady=5)
                date_frame.grid_propagate(False)
                
                # Configurar estilo del DateEntry
                style = ttk.Style()
                style.configure(
                    'Custom.DateEntry',
                    fieldbackground=EstiloApp.COLOR_PRINCIPAL,
                    background=EstiloApp.COLOR_PRINCIPAL,
                    foreground='black',
                    arrowcolor='black',
                    borderwidth=0,
                    highlightthickness=0,
                    relief='flat'
                )
                
                widget = DateEntry(
                    date_frame,
                    width=12,
                    background=EstiloApp.COLOR_PRINCIPAL,
                    foreground='black',
                    borderwidth=0,
                    font=('Roboto', 14),
                    date_pattern='dd-mm-yyyy',
                    style='Custom.DateEntry',
                    justify='center',
                    locale='es',
                    relief='flat'
                )
                widget.pack(expand=True, fill='both', padx=5, pady=2)
                
                # Eliminar bordes adicionales
                widget._top_cal.configure(relief='flat', borderwidth=0)
                for child in widget.winfo_children():
                    if isinstance(child, tk.Entry):
                        child.configure(relief='flat', borderwidth=0, highlightthickness=0)
                
                # Agregar checkbox "En curso" solo para el campo de fecha de alta
                if field[0] == "Fecha Alta:":
                    # Crear un frame adicional para contener el checkbox
                    check_frame = ctk.CTkFrame(
                        form_container,
                        fg_color="transparent"
                    )
                    check_frame.grid(row=field[2], column=field[3]+0, sticky="w", padx=(0, 105), pady=5)
                    
                    # Variable para controlar el estado del checkbox
                    self.var_en_curso = tk.BooleanVar(value=False)
                    
                    # Checkbox para indicar si el accidente está en curso
                    self.check_en_curso = ctk.CTkCheckBox(
                        check_frame,
                        text="En curso",
                        variable=self.var_en_curso,
                        command=self._toggle_fecha_alta
                    )
                    self.check_en_curso.pack(side="left", padx=5)
                    
            elif len(field) > 4 and field[4] == "text":
                widget = ctk.CTkTextbox(
                    form_container,
                    height=80,
                    width=400,
                    font=('Roboto', 14),
                    border_width=2,
                    border_color=EstiloApp.COLOR_SECUNDARIO,
                    corner_radius=8
                )
                widget.grid(row=field[2], column=field[3]+1, rowspan=field[5], 
                          sticky="nsew", padx=20, pady=5)
                setattr(self, field[1], widget)
                continue
            else:
                widget = ctk.CTkEntry(
                    form_container,
                    height=35,
                    width=200,
                    font=('Roboto', 14),
                    justify='center',
                    border_width=2,
                    border_color=EstiloApp.COLOR_SECUNDARIO,
                    corner_radius=8
                )
                widget.grid(row=field[2], column=field[3]+1, sticky="w", padx=20, pady=5)

            setattr(self, field[1], widget)

            # Vincular eventos al campo legajo
            if field[1] == "entry_legajo":
                widget.bind('<FocusOut>', self.buscar_empleado)
                widget.bind('<Return>', self.buscar_empleado)
    
    def _create_crud_buttons(self, parent):
        """Crear botones CRUD"""
        buttons_frame = ctk.CTkFrame(parent, fg_color="transparent")
        buttons_frame.grid(row=1, column=2, sticky="ne", padx=20, pady=20)

        buttons = [
            ("Insertar", self.guardar_registro, "#2ECC71", "#27AE60"),  # Verde
            ("Modificar", self.modificar_licencia, "#3498DB", "#2980B9"),  # Azul
            ("Eliminar", self._eliminar_licencia, "#E74C3C", "#C0392B"),  # Rojo
            ("Limpiar Todo", self.limpiar_campos, "#95A5A6", "#7F8C8D")  # Gris
        ]

        for idx, (text, command, color, hover_color) in enumerate(buttons):
            ctk.CTkButton(
                buttons_frame,
                text=text,
                command=command,
                font=ctk.CTkFont(size=14),
                fg_color=color,
                hover_color=hover_color,
                height=35,
                width=120,
                text_color="white"
            ).grid(row=idx, column=0, pady=5)
    
    def _create_table(self):
        """Crear tabla de accidentes"""
        table_frame = ctk.CTkFrame(
            self.main_frame,
            fg_color=EstiloApp.COLOR_FRAMES,
            corner_radius=10
        )
        table_frame.grid(row=1, column=0, sticky="nsew", padx=15, pady=(5, 15))
        
        title_label = ctk.CTkLabel(
            table_frame,
            text="Historial de Accidentes",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title_label.pack(pady=(15, 10))

        # Frame para la tabla con scrollbars
        tree_container = ctk.CTkFrame(table_frame, fg_color="transparent")
        tree_container.pack(fill="both", expand=True, padx=15, pady=(0, 15))

        # Crear Treeview con las columnas correctas
        self.tree = ttk.Treeview(
            tree_container,
            columns=("id_art", "legajo", "fecha_acc", "fecha_alta", "dias_baja", "diagnostico", "ambito", "objetivo", "n_siniestro", "descripcion"),
            show="headings",
            style="Custom.Treeview"
        )
        
        # Configurar scrollbars
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Configurar columnas
        columnas = {
            "id_art": {"texto": "ID", "ancho": 50},
            "legajo": {"texto": "Legajo", "ancho": 80},
            "fecha_acc": {"texto": "Fecha Accidente", "ancho": 120},
            "fecha_alta": {"texto": "Fecha Alta", "ancho": 120},
            "dias_baja": {"texto": "Días", "ancho": 60},
            "diagnostico": {"texto": "Diagnóstico", "ancho": 200},
            "descripcion": {"texto": "Descripción", "ancho": 200},
            "ambito": {"texto": "Ámbito", "ancho": 100},
            "objetivo": {"texto": "Descripción", "ancho": 200},
            "n_siniestro": {"texto": "N° Siniestro", "ancho": 100}
        }
        
        for col, config in columnas.items():
            self.tree.heading(col, text=config["texto"], anchor="center")
            self.tree.column(col, width=config["ancho"], anchor="center")

        # Ubicar componentes
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_container.grid_columnconfigure(0, weight=1)
        tree_container.grid_rowconfigure(0, weight=1)
        
        # Agregar el binding para doble clic y clic derecho
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        self.tree.bind("<Button-3>", self.mostrar_menu_contextual)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
    
    def _load_gif_frames(self):
        """Cargar frames del GIF del logo"""
        try:
            # Construir la ruta al archivo del logo
            logo_path = os.path.join(
                os.path.dirname(os.path.dirname(__file__)),
                'resources',
                'icons_gifs',
                'logo.gif'
            )
            
            if os.path.exists(logo_path):
                self.gif_frames = []
                self.current_frame = 0
                gif = Image.open(logo_path)
                
                # Calcular el tamaño manteniendo la proporción circular
                size = 300
                padding = 20  # Espacio alrededor del círculo
                circle_size = size - (padding * 2)
                
                for frame_index in range(0, gif.n_frames):
                    gif.seek(frame_index)
                    frame_image = gif.convert('RGBA')
                    
                    # Crear una imagen cuadrada con fondo transparente
                    square_image = Image.new('RGBA', (size, size), (0, 0, 0, 0))
                    
                    # Redimensionar el frame manteniendo la proporción
                    aspect_ratio = frame_image.width / frame_image.height
                    if aspect_ratio > 1:
                        new_width = circle_size
                        new_height = int(circle_size / aspect_ratio)
                    else:
                        new_height = circle_size
                        new_width = int(circle_size * aspect_ratio)
                    
                    frame_image = frame_image.resize((new_width, new_height), Image.LANCZOS)
                    
                    # Centrar la imagen en el cuadrado
                    x = (size - new_width) // 2
                    y = (size - new_height) // 2
                    square_image.paste(frame_image, (x, y))
                    
                    ctk_frame = ctk.CTkImage(
                        light_image=square_image,
                        dark_image=square_image,
                        size=(size, size)
                    )
                    self.gif_frames.append(ctk_frame)
                
                if self.gif_frames:
                    self.logo_label.configure(image=self.gif_frames[0])
                    self._animate_gif()
            else:
                self.logger.error(f"Archivo de logo no encontrado en: {logo_path}")
                self._create_placeholder_logo(self.logo_label)
                    
        except Exception as e:
            self.logger.error(f"Error cargando logo: {e}")
            self._create_placeholder_logo(self.logo_label)
    
    def _animate_gif(self):
        """Animar el GIF frame por frame usando CTkImage"""
        try:
            if (hasattr(self, 'logo_label') and 
                self.logo_label.winfo_exists() and 
                hasattr(self, 'gif_frames') and 
                self.gif_frames and 
                not self.is_destroyed):
                
                # Actualizar el frame actual
                self.logo_label.configure(image=self.gif_frames[self.current_frame])
                self.current_frame = (self.current_frame + 1) % len(self.gif_frames)
                
                # Programar siguiente frame
                if hasattr(self, 'main_container') and self.main_container.winfo_exists():
                    self.main_container.after(100, self._animate_gif)
            
        except Exception as e:
            self.logger.error(f"Error en animación del logo: {e}")
    
    def _create_placeholder_logo(self, parent):
        """Crear logo placeholder cuando no se puede cargar el GIF"""
        placeholder = ctk.CTkLabel(
            parent,
            text="LOGO",
            font=ctk.CTkFont(size=16, weight="bold"),
            width=50,
            height=50,
            fg_color=EstiloApp.COLOR_SECUNDARIO,
            corner_radius=10
        )
        placeholder.grid(row=0, column=0, rowspan=3, padx=(10, 20), pady=5)
    
    def show_in_frame(self, parent_frame):
        """Mostrar el módulo en un frame específico usando grid consistentemente"""
        try:
            # Limpiar el frame padre primero
            for widget in parent_frame.winfo_children():
                widget.destroy()
                
            # Configurar el frame padre
            parent_frame.grid_columnconfigure(0, weight=1)
            parent_frame.grid_rowconfigure(0, weight=1)
            
            # Actualizar referencias
            self.parent = parent_frame
            self.root = self._find_root_window(parent_frame)
            
            # Crear el contenedor principal usando grid
            self.main_container = ctk.CTkFrame(parent_frame, fg_color=EstiloApp.COLOR_PRINCIPAL)
            self.main_container.grid(row=0, column=0, sticky="nsew")
            
            # Configurar el grid del contenedor principal
            self.main_container.grid_columnconfigure(0, weight=1)
            self.main_container.grid_rowconfigure(0, weight=0)  # Header
            self.main_container.grid_rowconfigure(1, weight=0)  # Separator
            self.main_container.grid_rowconfigure(2, weight=1)  # Content
            
            # Inicializar la interfaz
            self.create_gui()
            
        except Exception as e:
            self.logger.error(f"Error en show_in_frame: {str(e)}")
            raise
    
    def _init_async(self):
        """
        Inicialización asíncrona de la aplicación
        """
        try:
            # Configurar la ventana
            self.setup_window()
            
            # Crear la interfaz
            self.create_gui()
            
            # Inicializar la base de datos
            self._init_database()
            
            # Ejecutar solo en modo standalone
            if self.is_standalone:
                # Iniciar la aplicación
                self.root.mainloop()
                
        except Exception as e:
            self.logger.error(f"Error en inicialización: {str(e)}")
            messagebox.showerror("Error", f"Error al inicializar la aplicación: {str(e)}")
    
    def _find_root_window(self, parent):
        """Encontrar la ventana raíz a partir de un widget padre"""
        if parent is None:
            return None
        
        widget = parent
        while widget.master is not None:
            widget = widget.master
        return widget
        
    def _init_database(self):
        """Inicializar la conexión a la base de datos"""
        try:
            # Asegurarse de que las variables de entorno estén cargadas
            load_dotenv()
            
            # Crear pool de conexiones - sin pasar argumentos al constructor
            # DatabasePool es un singleton que lee directamente de las variables de entorno
            self.db_pool = DatabasePool()
            
            # Probar conexión obteniendo una conexión
            try:
                conn = self.db_pool.get_connection()
                conn.close()  # Devolver la conexión al pool
                self.logger.info("Conexión a la base de datos establecida correctamente")
            except Exception as e:
                raise Exception(f"Error al probar la conexión: {str(e)}")
            
        except Exception as e:
            self.logger.error(f"Error al inicializar la base de datos: {str(e)}")
            if self.is_standalone:
                messagebox.showerror("Error de Conexión", 
                                    f"No se pudo conectar a la base de datos: {str(e)}")
                
    def create_scrollable_frame(self, parent):
        """Crear un frame con scroll"""
        # Crear frame contenedor
        container = ctk.CTkFrame(parent, fg_color="transparent")
        container.pack(fill="both", expand=True)
        
        # Crear canvas para el scroll
        canvas = tk.Canvas(
            container,
            bg=EstiloApp.COLOR_PRINCIPAL,
            highlightthickness=0
        )
        
        # Scrollbar vertical
        scrollbar = ttk.Scrollbar(
            container,
            orient="vertical",
            command=canvas.yview
        )
        
        # Frame interior
        scrollable_frame = ctk.CTkFrame(
            canvas,
            fg_color=EstiloApp.COLOR_PRINCIPAL,
            corner_radius=0
        )
        
        # Configurar el frame para expandirse
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        # Crear ventana en el canvas con ancho fijo
        canvas_window = canvas.create_window(
            (0, 0),
            window=scrollable_frame,
            anchor="nw"
        )
        
        # Ajustar el ancho del frame al canvas
        def adjust_window_size(event):
            canvas_width = event.width
            canvas.itemconfig(canvas_window, width=canvas_width)
        
        canvas.bind("<Configure>", adjust_window_size)
        
        # Posicionar elementos
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Configurar canvas
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Configurar scroll con rueda del ratón
        scrollable_frame.bind(
            "<Enter>",
            lambda e: canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        )
        scrollable_frame.bind(
            "<Leave>",
            lambda e: canvas.unbind_all("<MouseWheel>")
        )
        
        return scrollable_frame
        
    def on_closing(self):
        """Manejar el cierre de la ventana"""
        self.is_destroyed = True
        if self.is_standalone:
            self.root.destroy()
            
    def run(self):
        """Iniciar la aplicación"""
        if self.is_standalone:
            self.root.mainloop()
            
    def _create_photo_canvas(self, parent):
        """Crear canvas para mostrar la foto del empleado"""
        # Frame contenedor para la foto
        photo_frame = ctk.CTkFrame(
            parent,
            fg_color=EstiloApp.COLOR_FRAMES,
            corner_radius=10,
            width=150,
            height=180
        )
        photo_frame.grid(row=0, column=1, padx=15, pady=15, sticky="ne")
        photo_frame.grid_propagate(False)  # Mantener tamaño fijo
        
        # Canvas para la foto
        self.photo_canvas = tk.Canvas(
            photo_frame,
            width=130,
            height=160,
            bg=EstiloApp.COLOR_FRAMES,
            highlightthickness=0
        )
        self.photo_canvas.pack(expand=True, fill="both", padx=10, pady=10)
        
        # Mostrar placeholder inicialmente
        self._mostrar_placeholder_foto()
        
        return photo_frame
        
    def buscar_empleado(self, event=None):
        """Buscar empleado por legajo"""
        try:
            # Evitar múltiples consultas simultáneas
            if hasattr(self, '_consulta_en_progreso') and self._consulta_en_progreso:
                return
                
            self._consulta_en_progreso = True
            
            # Obtener legajo
            legajo = self.entry_legajo.get().strip()
            
            # Validar legajo
            if not legajo:
                self.mostrar_mensaje("Advertencia", "Debe ingresar un número de legajo", "warning")
                self._consulta_en_progreso = False
                return
                
            try:
                legajo = int(legajo)
            except ValueError:
                self.mostrar_mensaje("Error", "El legajo debe ser un número", "error")
                self._consulta_en_progreso = False
                return
                
            # Iniciar thread manager si no existe
            if not hasattr(self, 'thread_manager') or self.thread_manager is None:
                self.thread_manager = ThreadManager()
                
            # Mostrar indicador de carga
            self._mostrar_placeholder_foto()
            
            # Ejecutar consulta en segundo plano
            self.thread_manager.submit_task(
                "consulta_empleado",
                lambda: self._consultar_empleado_db(legajo),
                self._actualizar_ui_empleado
            )
        except Exception as e:
            self.logger.error(f"Error al buscar empleado: {str(e)}")
            self.mostrar_mensaje("Error", f"Error al buscar empleado: {str(e)}", "error")
            if hasattr(self, '_consulta_en_progreso'):
                self._consulta_en_progreso = False

    def _consultar_empleado_db(self, legajo):
        """Consultar datos del empleado en la base de datos"""
        conn = None
        cursor = None
        try:
            conn = self.db_pool.get_connection()
            cursor = conn.cursor()
            
            # Consultar datos del empleado
            query = """
            SELECT apellido_nombre, foto FROM personal WHERE legajo = %s
            """
            
            cursor.execute(query, (legajo,))
            empleado = cursor.fetchone()
            
            if not empleado:
                return None  # Empleado no encontrado
                
            apellido_nombre, foto_blob = empleado
            
            # Contar accidentes
            query_count = """
            SELECT COUNT(*) FROM accidentes WHERE legajo = %s
            """
            
            cursor.execute(query_count, (legajo,))
            total_accidentes = cursor.fetchone()[0]
            
            # Consultar accidentes
            self._consultar_accidentes(cursor, legajo)
            
            # Actualizar estadísticas
            self._actualizar_estadisticas(cursor, legajo)
            
            return (apellido_nombre, foto_blob, total_accidentes)
        except Exception as e:
            self.logger.error(f"Error al consultar empleado: {str(e)}")
            return e  # Devolver la excepción para manejarla en el callback
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

    def _actualizar_ui_empleado(self, result):
        """Actualizar UI con los resultados de la consulta de empleado"""
        try:
            # Liberar la bandera de consulta en progreso
            if hasattr(self, '_consulta_en_progreso'):
                self._consulta_en_progreso = False
                
            if result is None or result is False:
                self.mostrar_mensaje("Información", "Empleado no encontrado", "info")
                self._limpiar_formulario()
                return
                
            if isinstance(result, Exception):
                self.logger.error(f"Error en consulta de empleado: {str(result)}")
                self.handle_database_error(result, "consulta de empleado")
                return
                
            # Desempaquetar resultados
            apellido_nombre, foto_blob, total_accidentes = result
            
            # Actualizar datos en la UI
            def actualizar_ui():
                self._actualizar_datos_empleado(apellido_nombre, foto_blob, total_accidentes)
                
            self.root.after(0, actualizar_ui)
        except Exception as e:
            self.logger.error(f"Error al actualizar UI de empleado: {str(e)}")
            self.mostrar_mensaje("Error", f"Error al actualizar UI: {str(e)}", "error")

    def _actualizar_datos_empleado(self, apellido_nombre, foto_blob, total_accidentes):
        """Actualizar datos del empleado en la interfaz"""
        try:
            # Actualizar nombre
            if hasattr(self, 'nombre_empleado_label'):
                self.nombre_empleado_label.configure(text=f"Nombre: {apellido_nombre}")
            
            # Actualizar contador de accidentes
            if hasattr(self, 'total_accidentes_label'):
                self.total_accidentes_label.configure(text=f"Total Accidentes: {total_accidentes}")
            
            # Actualizar foto
            if foto_blob:
                try:
                    # Convertir blob a imagen
                    image_data = BytesIO(foto_blob)
                    pil_image = Image.open(image_data)
                    
                    # Redimensionar manteniendo proporción
                    width, height = pil_image.size
                    max_size = 130
                    
                    if width > height:
                        new_width = max_size
                        new_height = int(height * (max_size / width))
                    else:
                        new_height = max_size
                        new_width = int(width * (max_size / height))
                    
                    pil_image = pil_image.resize((new_width, new_height), Image.LANCZOS)
                    
                    # Convertir a PhotoImage
                    self.photo_image = ImageTk.PhotoImage(pil_image)
                    
                    # Limpiar canvas y mostrar imagen
                    self.photo_canvas.delete("all")
                    
                    # Calcular posición para centrar la imagen
                    canvas_width = self.photo_canvas.winfo_width()
                    canvas_height = self.photo_canvas.winfo_height()
                    x_position = (canvas_width - new_width) // 2
                    y_position = (canvas_height - new_height) // 2
                    
                    # Mostrar imagen centrada
                    self.photo_canvas.create_image(
                        x_position, y_position, 
                        anchor="nw", 
                        image=self.photo_image
                    )
                    
                    self.logger.info(f"Foto actualizada para {apellido_nombre}")
                except Exception as e:
                    self.logger.error(f"Error al procesar foto: {str(e)}")
                    self._mostrar_placeholder_foto()
            else:
                self._mostrar_placeholder_foto()
                
        except Exception as e:
            self.logger.error(f"Error al actualizar datos del empleado: {str(e)}")
            raise

    def _mostrar_placeholder_foto(self):
        """Mostrar imagen placeholder cuando no hay foto disponible"""
        try:
            # Limpiar canvas
            self.photo_canvas.delete("all")
            
            # Obtener dimensiones del canvas
            canvas_width = self.photo_canvas.winfo_width()
            canvas_height = self.photo_canvas.winfo_height()
            
            if canvas_width <= 1:  # Si el canvas aún no se ha dibujado
                canvas_width = 130
                canvas_height = 160
            
            # Crear rectángulo con bordes redondeados como fondo
            self.photo_canvas.create_rectangle(
                10, 10, canvas_width-10, canvas_height-10,
                fill="#E0E0E0",
                outline="#BBBBBB",
                width=2,
                stipple="gray50"
            )
            
            # Texto de placeholder
            self.photo_canvas.create_text(
                canvas_width//2, canvas_height//2,
                text="Sin\nFoto",
                font=("Roboto", 14, "bold"),
                fill="#757575",
                justify="center"
            )
            
        except Exception as e:
            self.logger.error(f"Error al mostrar placeholder de foto: {str(e)}")

    def _consultar_accidentes(self, cursor, legajo):
        """Consultar accidentes del empleado"""
        try:
            query = """
            SELECT id_art, legajo, fecha_acc, fecha_alta, dx, ambito, objetivo, n_siniestro, descripcion
            FROM accidentes 
            WHERE legajo = %s 
            ORDER BY fecha_acc DESC
            """
            
            cursor.execute(query, (legajo,))
            accidentes = []
            
            columns = [col[0] for col in cursor.description]
            
            for row in cursor.fetchall():
                # Convertir la tupla a un diccionario usando los nombres de columnas
                accidente = {}
                for i, value in enumerate(row):
                    if i < len(columns):
                        accidente[columns[i]] = value
                        
                # Asegurarse de que 'descripcion' exista
                if 'descripcion' not in accidente:
                    accidente['descripcion'] = ''
                    
                accidentes.append(accidente)
            
            def actualizar_treeview():
                try:
                    self._update_treeview(accidentes)
                except Exception as e:
                    self.logger.error(f"Error al actualizar treeview: {str(e)}")
                    # No mostrar mensaje aquí para evitar múltiples ventanas
            
            self.root.after(0, actualizar_treeview)
            
            return accidentes
        except Exception as e:
            self.logger.error(f"Error al consultar accidentes: {str(e)}")
            # No lanzar excepción para evitar interrumpir el flujo principal
            return []

    def _update_treeview(self, registros):
        """Actualizar el treeview con los registros obtenidos"""
        try:
            # Limpiar treeview
            self._clear_treeview()
            
            # Si no hay registros, salir
            if not registros:
                return
                
            # Verificar si el treeview tiene la columna 'descripcion'
            tiene_columna_descripcion = 'descripcion' in self.tree['columns']
            
            # Insertar nuevos registros
            for registro in registros:
                # Truncar textos largos para mejor visualización
                dx_truncado = registro['dx'][:50] + '...' if registro['dx'] and len(registro['dx']) > 50 else registro['dx']
                objetivo_truncado = registro['objetivo'][:50] + '...' if registro['objetivo'] and len(registro['objetivo']) > 50 else registro['objetivo']
                
                # Manejar la descripción - verificar si existe la clave
                descripcion = registro.get('descripcion', '')  # Usar get con valor predeterminado vacío
                descripcion_truncada = descripcion[:50] + '...' if descripcion and len(descripcion) > 50 else descripcion
                
                # Formatear fechas
                fecha_acc = registro['fecha_acc'].strftime('%d-%m-%Y') if registro['fecha_acc'] else ''
                fecha_alta = registro['fecha_alta'].strftime('%d-%m-%Y') if registro['fecha_alta'] else ''
                
                # Calcular días de baja correctamente
                dias_baja = 0
                if registro['fecha_acc'] and registro['fecha_alta']:
                    # Calcular manualmente los días entre las fechas
                    dias_baja = (registro['fecha_alta'] - registro['fecha_acc']).days
                elif registro['fecha_acc'] and not registro['fecha_alta']:
                    # Si no hay fecha de alta, calcular días hasta hoy
                    from datetime import datetime
                    dias_baja = (datetime.now().date() - registro['fecha_acc'].date()).days
                
                # Preparar valores base
                valores = [
                    registro['id_art'],
                    registro['legajo'],
                    fecha_acc,
                    fecha_alta,
                    dias_baja,  # Usar el valor calculado
                    dx_truncado,
                    registro['ambito'],
                    objetivo_truncado,
                    registro['n_siniestro']
                ]
                
                # Agregar descripción si la columna existe
                if tiene_columna_descripcion:
                    valores.append(descripcion_truncada)
                    
                # Insertar en el treeview
                self.tree.insert('', 'end', values=tuple(valores))
        except Exception as e:
            self.logger.error(f"Error al actualizar treeview: {str(e)}")
            # No mostrar mensaje aquí para evitar múltiples ventanas

    def _clear_treeview(self):
        """Limpiar todos los registros del treeview"""
        for item in self.tree.get_children():
            self.tree.delete(item)

    def _actualizar_estadisticas(self, cursor, legajo):
        """Actualizar estadísticas del empleado"""
        try:
            # Consultar último accidente
            query_ultimo = """
            SELECT fecha_acc, fecha_alta
            FROM accidentes 
            WHERE legajo = %s 
            ORDER BY fecha_acc DESC 
            LIMIT 1
            """
            
            cursor.execute(query_ultimo, (legajo,))
            ultimo = cursor.fetchone()
            
            # Consultar total de días de baja
            query_dias = """
            SELECT SUM(DATEDIFF(IFNULL(fecha_alta, CURDATE()), fecha_acc)) as total_dias
            FROM accidentes 
            WHERE legajo = %s
            """
            
            cursor.execute(query_dias, (legajo,))
            total_dias_result = cursor.fetchone()
            
            def _actualizar_labels():
                # Verificar si los labels existen antes de actualizarlos
                if hasattr(self, 'ultimo_accidente_label'):
                    if ultimo and ultimo[0]:  # Acceder por índice, no por clave
                        fecha_ultimo = ultimo[0].strftime('%d-%m-%Y') if ultimo[0] else "No registrado"
                        self.ultimo_accidente_label.configure(text=f"Último accidente: {fecha_ultimo}")
                    else:
                        self.ultimo_accidente_label.configure(text="Último accidente: No registrado")
                
                if hasattr(self, 'dias_totales_label'):
                    if total_dias_result and total_dias_result[0] is not None:  # Acceder por índice
                        total_dias = total_dias_result[0]
                        self.dias_totales_label.configure(text=f"Días totales de baja: {total_dias}")
                    else:
                        self.dias_totales_label.configure(text="Días totales de baja: 0")
        
            # Actualizar UI en el hilo principal
            self.root.after(0, _actualizar_labels)
        except Exception as e:
            self.logger.error(f"Error al actualizar estadísticas: {str(e)}")
            # No lanzar excepción para evitar interrumpir el flujo principal

    def on_tree_double_click(self, event):
        """Manejar doble clic en el treeview"""
        try:
            # Obtener el ítem seleccionado
            item = self.tree.selection()
            if not item:
                return
            
            # Obtener valores del ítem
            values = self.tree.item(item, 'values')
            if not values:
                return
            
            # Guardar ID seleccionado
            self.accidente_seleccionado_id = values[0]
            
            # Cargar datos en el formulario
            def actualizar_campos():
                self._cargar_datos_accidente(values)
            
            self.root.after(0, actualizar_campos)
        except Exception as e:
            self.logger.error(f"Error al cargar datos: {str(e)}")
            self.mostrar_mensaje("Error", f"No se pudieron cargar los datos: {str(e)}", "error")

    def _cargar_datos_accidente(self, values):
        """Cargar datos del accidente seleccionado en el formulario"""
        try:
            # Verificar que tenemos suficientes valores
            if len(values) < 9:  # Mínimo necesitamos 9 valores (sin descripción)
                self.logger.error(f"Datos insuficientes: {values}")
                return
            
            # Asignar valores a los campos
            # Legajo
            if hasattr(self, 'entry_legajo'):
                self.entry_legajo.delete(0, 'end')
                self.entry_legajo.insert(0, values[1])
            
            # Fecha de accidente
            if hasattr(self, 'entry_fecha_accidente'):
                fecha_acc = values[2]
                if fecha_acc:
                    try:
                        self.entry_fecha_accidente.set_date(datetime.strptime(fecha_acc, '%d-%m-%Y'))
                    except Exception as e:
                        self.logger.error(f"Error al establecer fecha de accidente: {str(e)}")
            
            # Fecha de alta
            if hasattr(self, 'entry_fecha_alta'):
                fecha_alta = values[3]
                if fecha_alta:
                    try:
                        self.entry_fecha_alta.set_date(datetime.strptime(fecha_alta, '%d-%m-%Y'))
                    except Exception as e:
                        self.logger.error(f"Error al establecer fecha de alta: {str(e)}")
            
            # Diagnóstico
            if hasattr(self, 'entry_diagnostico'):
                self.entry_diagnostico.delete(0, 'end')
                self.entry_diagnostico.insert(0, values[5])
            
            # Ámbito
            if hasattr(self, 'entry_ambito'):
                self.entry_ambito.delete(0, 'end')
                self.entry_ambito.insert(0, values[6])
            
            # Descripción - Usar el valor de descripción (values[9]) si está disponible
            if hasattr(self, 'text_descripcion'):
                self.text_descripcion.delete('1.0', 'end')
                # Si hay descripción (values[9]), usamos esa, sino usamos objetivo (values[7])
                descripcion = values[9] if len(values) > 9 else values[7]
                self.text_descripcion.insert('1.0', descripcion)
            
            # N° Siniestro
            if hasattr(self, 'entry_n_siniestro'):
                self.entry_n_siniestro.delete(0, 'end')
                self.entry_n_siniestro.insert(0, values[8])
            
            # Guardar ID seleccionado
            self.accidente_seleccionado_id = values[0]
            
            self.logger.info(f"Datos cargados correctamente para accidente ID: {values[0]}")
        except Exception as e:
            self.logger.error(f"Error al cargar datos: {str(e)}")
            self.mostrar_mensaje("Error", f"Error al cargar datos: {str(e)}", "error")

    def limpiar_campos_parcial(self):
        """Limpiar campos del formulario excepto legajo"""
        try:
            self.logger.info("Limpiando campos parcialmente...")
            
            # Limpiar fecha de accidente
            if hasattr(self, 'entry_fecha_accidente'):
                try:
                    self.entry_fecha_accidente.set_date(None)
                except Exception as e:
                    self.logger.warning(f"No se pudo limpiar fecha de accidente: {str(e)}")
            
            # Limpiar fecha de alta
            if hasattr(self, 'entry_fecha_alta'):
                try:
                    self.entry_fecha_alta.set_date(None)
                except Exception as e:
                    self.logger.warning(f"No se pudo limpiar fecha de alta: {str(e)}")
            
            # Limpiar diagnóstico
            if hasattr(self, 'entry_diagnostico'):
                self.entry_diagnostico.delete(0, 'end')
            
            # Limpiar ámbito
            if hasattr(self, 'entry_ambito'):
                self.entry_ambito.delete(0, 'end')
            
            # Limpiar descripción 
            if hasattr(self, 'text_descripcion'):
                self.text_descripcion.delete('1.0', 'end')
            
            # Limpiar N° Siniestro
            if hasattr(self, 'entry_n_siniestro'):
                self.entry_n_siniestro.delete(0, 'end')
            
            # Limpiar ID seleccionado
            self.accidente_seleccionado_id = None
            
            self.logger.info("Campos limpiados correctamente")
        except Exception as e:
            self.logger.error(f"Error al limpiar campos: {str(e)}")


    def _limpiar_formulario(self):
        """Limpiar todos los campos del formulario"""
        # Limpiar legajo
        self.entry_legajo.delete(0, tk.END)
        
        # Limpiar nombre
        self.nombre_completo_label.configure(text="👤 Empleado: ")
        
        # Limpiar foto
        self._mostrar_placeholder_foto()
        
        # Limpiar estadísticas
        self.total_accidentes_label.configure(text="📋 Historial: 0 accidentes registrados")
        self.ultimo_accidente_label.configure(text="⏱️ Último accidente: No registrado")
        self.dias_totales_label.configure(text="📊 Total días con ART: 0 días")
        
        # Limpiar campos específicos
        self.limpiar_campos_parcial()
        
        # Limpiar treeview
        self._clear_treeview()
        
        # Resetear variables
        self._ultimo_legajo_consultado = None

    def mostrar_mensaje(self, titulo, mensaje, tipo="info"):
        """Mostrar mensaje al usuario"""
        if tipo == "info":
            messagebox.showinfo(titulo, mensaje)
        elif tipo == "warning":
            messagebox.showwarning(titulo, mensaje)
        elif tipo == "error":
            messagebox.showerror(titulo, mensaje)

    def _mostrar_dialogo_confirmacion(self, titulo, mensaje, accion_confirmacion):
        """Mostrar diálogo de confirmación"""
        respuesta = messagebox.askyesno(titulo, mensaje)
        if respuesta:
            accion_confirmacion()

    def handle_database_error(self, error, operacion):
        """Manejar errores de base de datos"""
        self.logger.error(f"Error de base de datos en {operacion}: {str(error)}")
        self.mostrar_mensaje("Error de Base de Datos", 
                             f"Error al {operacion}: {str(error)}", 
                             "error")

    def guardar_registro(self):
        """Guardar un nuevo registro de accidente"""
        try:
            # Validar campos
            if not self.validar_campos():
                return
            
            # Obtener valores de los campos
            legajo = self.entry_legajo.get().strip()
            
            # Obtener fechas de los DateEntry
            fecha_acc = self.entry_fecha_accidente.get_date()
            fecha_alta = self.entry_fecha_alta.get_date()
            
            # Calcular días de baja
            dias_baja = (fecha_alta - fecha_acc).days
            
            # Obtener resto de campos
            dx = self.entry_diagnostico.get().strip()
            ambito = self.entry_ambito.get().strip()
            objetivo = self.text_descripcion.get("1.0", "end-1c").strip()
            n_siniestro = self.entry_n_siniestro.get().strip()
            
            # Ejecutar inserción en un hilo separado
            if not hasattr(self, 'thread_manager'):
                self.thread_manager = ThreadManager()
            
            # Pasar el tipo de tarea como primer argumento
            self.thread_manager.submit_task(
                "insertar_accidente",  # Tipo de tarea
                lambda: self._insertar_accidente_db(legajo, fecha_acc, fecha_alta, dias_baja, dx, ambito, objetivo, n_siniestro),
                self._actualizar_ui_insercion
            )
            
        except Exception as e:
            self.logger.error(f"Error al guardar registro: {str(e)}")
            self.mostrar_mensaje("Error", f"Error al guardar registro: {str(e)}", "error")

    def _insertar_accidente_db(self, legajo, fecha_acc, fecha_alta, dias_baja, dx, ambito, objetivo, n_siniestro):
        """Insertar accidente en la base de datos"""
        try:
            conn = self.db_pool.get_connection()
            cursor = conn.cursor()
            
            # Obtener descripción del campo text_descripcion
            descripcion = ""
            if hasattr(self, 'text_descripcion'):
                descripcion = self.text_descripcion.get('1.0', 'end-1c')
            
            query = """
            INSERT INTO accidentes (
                legajo, fecha_acc, fecha_alta, dx, ambito, objetivo, n_siniestro, descripcion
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """
            
            cursor.execute(query, (
                legajo, fecha_acc, fecha_alta, dx, ambito, objetivo, n_siniestro, descripcion
            ))
            
            conn.commit()
            
            # Obtener el ID del registro insertado
            accidente_id = cursor.lastrowid
            
            # Consultar el registro completo
            cursor.execute("""
            SELECT id_art, legajo, fecha_acc, fecha_alta, dx, ambito, objetivo, n_siniestro, descripcion
            FROM accidentes WHERE id_art = %s
            """, (accidente_id,))
            
            nuevo_registro = cursor.fetchone()
            
            # Convertir a diccionario para mantener consistencia
            if nuevo_registro:
                columns = [col[0] for col in cursor.description]
                result_dict = {}
                for i, value in enumerate(nuevo_registro):
                    if i < len(columns):
                        result_dict[columns[i]] = value
                
                # Calcular días de baja si no están en el resultado
                if 'fecha_acc' in result_dict and 'fecha_alta' in result_dict:
                    if result_dict['fecha_acc'] and result_dict['fecha_alta']:
                        result_dict['dias_baja'] = (result_dict['fecha_alta'] - result_dict['fecha_acc']).days
                    elif result_dict['fecha_acc']:
                        from datetime import datetime
                        result_dict['dias_baja'] = (datetime.now().date() - result_dict['fecha_acc'].date()).days
                
                return result_dict
            return True
        except Exception as e:
            self.logger.error(f"Error al insertar accidente: {str(e)}")
            raise
        finally:
            if 'cursor' in locals() and cursor:
                cursor.close()
            if 'conn' in locals() and conn:
                conn.close()

    def _actualizar_ui_insercion(self, result):
        """Actualizar la UI después de insertar un registro"""
        try:
            if result:
                self.mostrar_mensaje("Éxito", "Registro guardado correctamente", "info")
                
                # Actualizar treeview con el nuevo registro
                self.buscar_empleado()
                
                # Limpiar formulario
                self.limpiar_campos_parcial()
            else:
                self.mostrar_mensaje("Error", "Error desconocido al guardar el registro", "error")
                
        except Exception as e:
            self.logger.error(f"Error al actualizar UI después de inserción: {str(e)}")
            self.mostrar_mensaje("Error", f"Error al actualizar UI: {str(e)}", "error")

    def validar_campos(self):
        """Validar campos del formulario"""
        # Validar legajo
        if not hasattr(self, 'entry_legajo') or not self.entry_legajo.get().strip():
            self.mostrar_mensaje("Error", "Debe ingresar un número de legajo", "error")
            return False
        
        # Validar fechas
        try:
            if not hasattr(self, 'entry_fecha_accidente'):
                self.mostrar_mensaje("Error", "Campo de fecha de accidente no encontrado", "error")
                return False
            fecha_acc = self.entry_fecha_accidente.get_date()
        except Exception as e:
            self.logger.error(f"Error al obtener fecha de accidente: {str(e)}")
            self.mostrar_mensaje("Error", "Debe ingresar una fecha de accidente válida", "error")
            return False
        
        try:
            if not hasattr(self, 'entry_fecha_alta'):
                self.mostrar_mensaje("Error", "Campo de fecha de alta no encontrado", "error")
                return False
            fecha_alta = self.entry_fecha_alta.get_date()
        except Exception as e:
            self.logger.error(f"Error al obtener fecha de alta: {str(e)}")
            self.mostrar_mensaje("Error", "Debe ingresar una fecha de alta válida", "error")
            return False
        
        # Validar que la fecha de alta sea posterior a la fecha de accidente
        if fecha_alta < fecha_acc:
            self.mostrar_mensaje("Error", "La fecha de alta debe ser posterior a la fecha de accidente", "error")
            return False
        
        # Validar diagnóstico
        if not hasattr(self, 'entry_diagnostico') or not self.entry_diagnostico.get().strip():
            self.mostrar_mensaje("Error", "Debe ingresar un diagnóstico", "error")
            return False
        
        return True

    def modificar_licencia(self):
        """Modificar un registro de accidente existente"""
        try:
            # Verificar si hay un registro seleccionado
            if not hasattr(self, 'accidente_seleccionado_id') or not self.accidente_seleccionado_id:
                self.mostrar_mensaje("Advertencia", "Debe seleccionar un registro para modificar", "warning")
                return
            
            # Validar campos
            if not self.validar_campos():
                return
            
            # Confirmar modificación
            self._mostrar_dialogo_confirmacion(
                "Confirmar Modificación",
                "¿Está seguro de que desea modificar este registro?",
                self._ejecutar_modificacion
            )
            
        except Exception as e:
            self.logger.error(f"Error al iniciar modificación: {str(e)}")
            self.mostrar_mensaje("Error", f"Error al iniciar modificación: {str(e)}", "error")

    def _ejecutar_modificacion(self):
        """Ejecutar la modificación del accidente"""
        try:
            # Validar que hay un accidente seleccionado
            if not hasattr(self, 'accidente_seleccionado_id') or not self.accidente_seleccionado_id:
                self.mostrar_mensaje("Advertencia", "Debe seleccionar un accidente para modificar", "warning")
                return
                
            # Validar campos
            if not self.validar_campos():
                return
                
            # Obtener valores
            legajo = int(self.entry_legajo.get())
            fecha_acc = self.entry_fecha_accidente.get_date()
            fecha_alta = self.entry_fecha_alta.get_date()
            
            # Calcular días de baja
            dias_baja = None
            if fecha_acc and fecha_alta:
                dias_baja = (fecha_alta - fecha_acc).days
                
            dx = self.entry_diagnostico.get()
            ambito = self.entry_ambito.get()
            
            # Obtener descripción del campo text_descripcion
            objetivo = ""
            if hasattr(self, 'text_descripcion'):
                objetivo = self.text_descripcion.get('1.0', 'end-1c')
                
            n_siniestro = self.entry_n_siniestro.get()
            
            # Iniciar thread manager si no existe
            if not hasattr(self, 'thread_manager') or self.thread_manager is None:
                self.thread_manager = ThreadManager()
                
            # Ejecutar modificación en segundo plano
            self.thread_manager.submit_task(
                "modificar_accidente",  # Tipo de tarea
                lambda: self._modificar_accidente_db(
                    self.accidente_seleccionado_id, legajo, fecha_acc, fecha_alta, 
                    dias_baja, dx, ambito, objetivo, n_siniestro
                ),
                self._actualizar_ui_modificacion
            )
        except Exception as e:
            self.logger.error(f"Error al ejecutar modificación: {str(e)}")
            self.mostrar_mensaje("Error", f"Error al modificar registro: {str(e)}", "error")

    def _modificar_accidente_db(self, accidente_id, legajo, fecha_acc, fecha_alta, dias_baja, dx, ambito, objetivo, n_siniestro):
        """Modificar accidente en la base de datos"""
        try:
            conn = self.db_pool.get_connection()
            cursor = conn.cursor()
            
            # Usar el valor de objetivo como descripción
            descripcion = objetivo
            
            query = """
            UPDATE accidentes SET 
                legajo = %s, 
                fecha_acc = %s, 
                fecha_alta = %s, 
                dx = %s, 
                ambito = %s, 
                objetivo = %s, 
                n_siniestro = %s,
                descripcion = %s
            WHERE id_art = %s
            """
            
            cursor.execute(query, (
                legajo, fecha_acc, fecha_alta, dx, ambito, objetivo, n_siniestro, descripcion, accidente_id
            ))
            
            conn.commit()
            
            # Verificar si se modificó correctamente
            if cursor.rowcount > 0:
                # Consultar el registro actualizado
                cursor.execute("""
                SELECT id_art, legajo, fecha_acc, fecha_alta, dx, ambito, objetivo, n_siniestro, descripcion
                FROM accidentes WHERE id_art = %s
                """, (accidente_id,))
                
                registro_actualizado = cursor.fetchone()
                
                # Convertir a diccionario para mantener consistencia
                if registro_actualizado:
                    columns = [col[0] for col in cursor.description]
                    result_dict = {}
                    for i, value in enumerate(registro_actualizado):
                        if i < len(columns):
                            result_dict[columns[i]] = value
                    
                    # Calcular días de baja si no están en el resultado
                    if 'fecha_acc' in result_dict and 'fecha_alta' in result_dict:
                        if result_dict['fecha_acc'] and result_dict['fecha_alta']:
                            result_dict['dias_baja'] = (result_dict['fecha_alta'] - result_dict['fecha_acc']).days
                        elif result_dict['fecha_acc']:
                            from datetime import datetime
                            result_dict['dias_baja'] = (datetime.now().date() - result_dict['fecha_acc'].date()).days
                    
                    return result_dict
                return True
            else:
                return False
        except Exception as e:
            self.logger.error(f"Error al modificar accidente: {str(e)}")
            raise
        finally:
            if 'cursor' in locals() and cursor:
                cursor.close()
            if 'conn' in locals() and conn:
                conn.close()

    def _actualizar_ui_modificacion(self, result):
        """Actualizar UI después de modificar un accidente"""
        try:
            if isinstance(result, bool):
                if result:
                    self.mostrar_mensaje("Éxito", "Registro modificado correctamente", "info")
                    # Actualizar la vista
                    self.buscar_empleado()
                    # Limpiar campos
                    self.limpiar_campos_parcial()
                else:
                    self.mostrar_mensaje("Error", "No se pudo modificar el registro", "error")
                return
                
            if isinstance(result, Exception):
                self.logger.error(f"Error al modificar accidente: {str(result)}")
                self.handle_database_error(result, "modificación de accidente")
                return
                
            if isinstance(result, dict):
                self.mostrar_mensaje("Éxito", "Registro modificado correctamente", "info")
                # Actualizar la vista
                self.buscar_empleado()
                # Limpiar campos
                self.limpiar_campos_parcial()
            elif isinstance(result, tuple) and len(result) > 0:
                self.mostrar_mensaje("Éxito", "Registro modificado correctamente", "info")
                # Actualizar la vista
                self.buscar_empleado()
                # Limpiar campos
                self.limpiar_campos_parcial()
            else:
                self.mostrar_mensaje("Error", "No se pudo modificar el registro", "error")
        except Exception as e:
            self.logger.error(f"Error al actualizar UI después de modificación: {str(e)}")
            self.mostrar_mensaje("Error", f"Error al actualizar UI: {str(e)}", "error")

    def _eliminar_licencia(self):
        """Eliminar registro de accidente"""
        # Verificar si hay un registro seleccionado
        if not hasattr(self, 'accidente_seleccionado_id') or not self.accidente_seleccionado_id:
            self.mostrar_mensaje("Error", "Debe seleccionar un registro para eliminar", "error")
            return
        
        # Confirmar eliminación
        self._mostrar_dialogo_confirmacion(
            "Confirmar Eliminación",
            "¿Está seguro de eliminar este registro? Esta acción no se puede deshacer.",
            self._ejecutar_eliminacion
        )

    def _ejecutar_eliminacion(self):
        """Ejecutar la eliminación del registro"""
        try:
            # Ejecutar eliminación en un hilo separado
            if not hasattr(self, 'thread_manager'):
                self.thread_manager = ThreadManager()
            
            # Usar una función lambda para encapsular la llamada y solo pasar el resultado al callback
            accidente_id = self.accidente_seleccionado_id
            self.thread_manager.submit_task(
                "eliminar_accidente",  # Tipo de tarea
                lambda: self._eliminar_accidente_db(accidente_id),
                self._actualizar_ui_eliminacion
            )
            
        except Exception as e:
            self.logger.error(f"Error al ejecutar eliminación: {str(e)}")
            self.mostrar_mensaje("Error", f"Error al ejecutar eliminación: {str(e)}", "error")

    def _eliminar_accidente_db(self, accidente_id):
        """Eliminar accidente de la base de datos"""
        try:
            conn = self.db_pool.get_connection()
            cursor = conn.cursor()
            
            query = "DELETE FROM accidentes WHERE id_art = %s"
            cursor.execute(query, (accidente_id,))
            
            conn.commit()
            
            return True
        except Exception as e:
            self.logger.error(f"Error al eliminar accidente: {str(e)}")
            raise
        finally:
            if 'cursor' in locals() and cursor:
                cursor.close()
            if 'conn' in locals() and conn:
                conn.close()

    def _actualizar_ui_eliminacion(self, result):
        """Actualizar la UI después de eliminar un registro"""
        try:
            # Verificar si el resultado es un booleano (éxito/fracaso simple)
            if isinstance(result, bool):
                if result:
                    self.mostrar_mensaje("Éxito", "Registro eliminado correctamente", "info")
                    # Actualizar treeview
                    self.buscar_empleado()
                    # Limpiar formulario
                    self.limpiar_campos_parcial()
                    # Resetear accidente seleccionado
                    self.accidente_seleccionado_id = None
                else:
                    self.mostrar_mensaje("Error", "Error al eliminar el registro", "error")
                return
                
            # Si no es booleano, verificar si es un diccionario o tupla con datos
            if result:
                self.mostrar_mensaje("Éxito", "Registro eliminado correctamente", "info")
                
                # Actualizar treeview sin el registro eliminado
                self.buscar_empleado()
                
                # Limpiar formulario
                self.limpiar_campos_parcial()
                
                # Resetear accidente seleccionado
                self.accidente_seleccionado_id = None
            else:
                self.mostrar_mensaje("Error", "Error desconocido al eliminar el registro", "error")
                
        except Exception as e:
            self.logger.error(f"Error al actualizar UI después de eliminación: {str(e)}")
            self.mostrar_mensaje("Error", f"Error al actualizar UI: {str(e)}", "error")

    def mostrar_menu_contextual(self, event):
        """Mostrar menú contextual en el treeview"""
        # Seleccionar ítem bajo el cursor
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            
            # Crear menú contextual
            menu = tk.Menu(self.root, tearoff=0)
            menu.add_command(label="Editar", command=lambda: self.on_tree_double_click(None))
            menu.add_command(label="Eliminar", command=self._eliminar_licencia)
            menu.add_separator()
            menu.add_command(label="Ver detalles completos", command=lambda: self._mostrar_detalles_completos(item))
            
            # Mostrar menú en la posición del cursor
            menu.post(event.x_root, event.y_root)

    def _mostrar_detalles_completos(self, item):
        """Mostrar ventana con detalles completos del registro"""
        values = self.tree.item(item, 'values')
        if not values:
            return
        
        # Crear ventana de detalles
        detalle_window = tk.Toplevel(self.root)
        detalle_window.title("Detalles del Accidente")
        detalle_window.geometry("600x500")
        detalle_window.grab_set()  # Modal
        
        # Configurar estilo
        detalle_frame = ctk.CTkFrame(detalle_window)
        detalle_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título
        titulo = ctk.CTkLabel(
            detalle_frame,
            text=f"Detalles del Accidente - Legajo {values[1]}",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        titulo.pack(pady=(0, 20))
        
        # Crear grid para los detalles
        info_frame = ctk.CTkFrame(detalle_frame, fg_color="transparent")
        info_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Etiquetas y valores
        labels = [
            "ID:", "Legajo:", "Fecha de Accidente:", "Fecha de Alta:",
            "Días de Baja:", "Diagnóstico:", "Ámbito:", "Objetivo:", "N° Siniestro:", "Descripción:"
        ]
        
        for i, label in enumerate(labels):
            lbl = ctk.CTkLabel(
                info_frame,
                text=label,
                font=ctk.CTkFont(size=12, weight="bold"),
                anchor="w"
            )
            lbl.grid(row=i, column=0, sticky="w", padx=5, pady=5)
            
            # Para diagnóstico y objetivo, usar text box
            if label in ["Diagnóstico:", "Objetivo:", "Descripción:"]:
                txt = ctk.CTkTextbox(
                    info_frame,
                    font=ctk.CTkFont(size=12),
                    height=60,
                    width=400
                )
                txt.grid(row=i, column=1, sticky="ew", padx=5, pady=5)
                txt.insert("1.0", values[i])
                txt.configure(state="disabled")
            else:
                val = ctk.CTkLabel(
                    info_frame,
                    text=values[i] if i < len(values) else "",
                    font=ctk.CTkFont(size=12),
                    anchor="w"
                )
                val.grid(row=i, column=1, sticky="w", padx=5, pady=5)
        
        # Botón cerrar
        cerrar_btn = ctk.CTkButton(
            detalle_frame,
            text="Cerrar",
            command=detalle_window.destroy
        )
        cerrar_btn.pack(pady=20)

    def on_tree_select(self, event):
        """Manejar selección en el treeview"""
        # Obtener el ítem seleccionado
        item = self.tree.selection()
        if not item:
            return
        
        # Obtener valores del ítem
        values = self.tree.item(item, 'values')
        if not values:
            return
        
        # Guardar ID seleccionado
        self.accidente_seleccionado_id = values[0]

    def exportar_a_excel(self):
        """Exportar datos a Excel"""
        # Verificar si hay datos para exportar
        if not self.tree.get_children():
            self.mostrar_mensaje("Información", "No hay datos para exportar", "info")
            return
        
        # Solicitar ubicación para guardar
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Guardar como"
        )
        
        if not file_path:
            return  # Usuario canceló
        
        # Ejecutar exportación en un hilo separado
        self.thread_manager.run_in_thread(
            target=self._exportar_datos_excel,
            args=(file_path,),
            callback=self._actualizar_ui_exportacion
        )

    def _exportar_datos_excel(self, file_path):
        """Exportar datos a Excel en segundo plano"""
        try:
            # Obtener datos del treeview
            data = []
            columns = [
                "ID", "Legajo", "Fecha Accidente", "Fecha Alta", 
                "Días Baja", "Diagnóstico", "Ámbito", "Objetivo", "N° Siniestro", "Descripción"
            ]
            
            for item in self.tree.get_children():
                values = self.tree.item(item, 'values')
                data.append(list(values))
            
            # Crear DataFrame
            df = pd.DataFrame(data, columns=columns)
            
            # Guardar a Excel
            df.to_excel(file_path, index=False, sheet_name="Accidentes")
            
            # Aplicar formato al archivo Excel
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            
            # Cargar el archivo guardado
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Dar formato a los encabezados
            header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar anchos de columna
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar con formato
            wb.save(file_path)
            
            return file_path
        except Exception as e:
            self.logger.error(f"Error al exportar a Excel: {str(e)}")
            raise

    def _actualizar_ui_exportacion(self, result):
        """Actualizar UI después de exportación"""
        if isinstance(result, Exception):
            self.mostrar_mensaje("Error", f"Error al exportar datos: {str(result)}", "error")
            return
        
        self.mostrar_mensaje("Éxito", f"Datos exportados correctamente a:\n{result}", "info")

    def generar_informe(self):
        """Generar informe de accidentes"""
        # Verificar si hay un legajo consultado
        if not self._ultimo_legajo_consultado:
            self.mostrar_mensaje("Información", "Primero debe consultar un empleado", "info")
            return
        
        # Ejecutar generación en un hilo separado
        # Cambiado run_in_thread por submit_task
        self.thread_manager.submit_task(
            "generar_informe",  # Tipo de tarea
            self._generar_informe_db,  # Función a ejecutar
            self._mostrar_informe,  # Callback
            self._ultimo_legajo_consultado  # Argumentos
        )

    def _generar_informe_db(self, legajo):
        """Generar datos para el informe desde la base de datos"""
        try:
            conn = self.db_pool.get_connection()
            cursor = conn.cursor(dictionary=True)
            
            # Consultar datos del empleado
            query_empleado = """
            SELECT p.legajo, p.apellido_nombre, p.foto, p.puesto, p.sector
            FROM personal p
            WHERE p.legajo = %s
            """
            cursor.execute(query_empleado, (legajo,))
            empleado = cursor.fetchone()
            
            if not empleado:
                return None, None, 0
            
            # Consultar historial de accidentes
            query_accidentes = """
            SELECT COUNT(*) as total_accidentes
            FROM accidentes
            WHERE legajo = %s
            """
            cursor.execute(query_accidentes, (legajo,))
            result = cursor.fetchone()
            total_accidentes = result['total_accidentes'] if result else 0
            
            # Consultar estadísticas adicionales
            self._actualizar_estadisticas(cursor, legajo)
            
            # Consultar accidentes del empleado
            self._consultar_accidentes(cursor, legajo)
            
            return empleado['apellido_nombre'], empleado['foto'], total_accidentes
        except Exception as e:
            self.logger.error(f"Error en consulta de empleado: {str(e)}")
            raise
        finally:
            if 'cursor' in locals() and cursor:
                cursor.close()
            if 'conn' in locals() and conn:
                conn.close()

    def _mostrar_informe(self, result):
        """Mostrar ventana con el informe generado"""
        if isinstance(result, Exception):
            self.mostrar_mensaje("Error", f"Error al generar informe: {str(result)}", "error")
            return
        
        if not result:
            self.mostrar_mensaje("Información", "No se encontraron datos para el informe", "info")
            return
        
        # Crear ventana de informe
        informe_window = tk.Toplevel(self.root)
        informe_window.title(f"Informe de Accidentes - Legajo {result['empleado']['legajo']}")
        informe_window.geometry("800x600")
        informe_window.grab_set()  # Modal
        
        # Frame principal
        main_frame = ctk.CTkFrame(informe_window)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Crear canvas para scroll
        canvas = tk.Canvas(
            main_frame,
            bg=EstiloApp.COLOR_PRINCIPAL,
            highlightthickness=0
        )
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        
        # Frame para contenido
        content_frame = ctk.CTkFrame(canvas, fg_color=EstiloApp.COLOR_PRINCIPAL)
        canvas.create_window((0, 0), window=content_frame, anchor="nw", width=canvas.winfo_reqwidth())
        
        # Configurar scroll
        def _configure_canvas(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        content_frame.bind("<Configure>", _configure_canvas)
        
        # Título del informe
        titulo = ctk.CTkLabel(
            content_frame,
            text=f"INFORME DE ACCIDENTES",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        titulo.pack(pady=(20, 10))
        
        # Datos del empleado
        empleado_frame = ctk.CTkFrame(content_frame, fg_color=EstiloApp.COLOR_FRAMES)
        empleado_frame.pack(fill="x", padx=20, pady=10)
        
        # Título sección empleado
        titulo_empleado = ctk.CTkLabel(
            empleado_frame,
            text="DATOS DEL EMPLEADO",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        titulo_empleado.pack(pady=10)
        
        # Grid para datos del empleado
        datos_frame = ctk.CTkFrame(empleado_frame, fg_color="transparent")
        datos_frame.pack(fill="x", padx=20, pady=10)
        
        # Etiquetas y valores
        ctk.CTkLabel(datos_frame, text="Legajo:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, sticky="w", padx=5, pady=5)
        ctk.CTkLabel(datos_frame, text=result['empleado']['legajo']).grid(row=0, column=1, sticky="w", padx=5, pady=5)
        
        ctk.CTkLabel(datos_frame, text="Nombre:", font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, sticky="w", padx=5, pady=5)
        ctk.CTkLabel(datos_frame, text=result['empleado']['apellido_nombre']).grid(row=1, column=1, sticky="w", padx=5, pady=5)
        
        ctk.CTkLabel(datos_frame, text="Puesto:", font=ctk.CTkFont(weight="bold")).grid(row=2, column=0, sticky="w", padx=5, pady=5)
        ctk.CTkLabel(datos_frame, text=result['empleado'].get('puesto', 'No especificado')).grid(row=2, column=1, sticky="w", padx=5, pady=5)
        
        ctk.CTkLabel(datos_frame, text="Sector:", font=ctk.CTkFont(weight="bold")).grid(row=3, column=0, sticky="w", padx=5, pady=5)
        ctk.CTkLabel(datos_frame, text=result['empleado'].get('sector', 'No especificado')).grid(row=3, column=1, sticky="w", padx=5, pady=5)
        
        # Estadísticas
        stats_frame = ctk.CTkFrame(content_frame, fg_color=EstiloApp.COLOR_FRAMES)
        stats_frame.pack(fill="x", padx=20, pady=10)
        
        # Título sección estadísticas
        titulo_stats = ctk.CTkLabel(
            stats_frame,
            text="ESTADÍSTICAS",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        titulo_stats.pack(pady=10)
        
        # Grid para estadísticas
        stats_grid = ctk.CTkFrame(stats_frame, fg_color="transparent")
        stats_grid.pack(fill="x", padx=20, pady=10)
        
        # Formatear fecha del último accidente
        ultimo_acc = "No registrado"
        if result['estadisticas']['ultimo_accidente']:
            ultimo_acc = result['estadisticas']['ultimo_accidente'].strftime('%d-%m-%Y')
        
        # Etiquetas y valores
        ctk.CTkLabel(stats_grid, text="Total Accidentes:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, sticky="w", padx=5, pady=5)
        ctk.CTkLabel(stats_grid, text=str(result['estadisticas']['total_accidentes'])).grid(row=0, column=1, sticky="w", padx=5, pady=5)
        
        ctk.CTkLabel(stats_grid, text="Total Días de Baja:", font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, sticky="w", padx=5, pady=5)
        ctk.CTkLabel(stats_grid, text=str(result['estadisticas']['total_dias'] or 0)).grid(row=1, column=1, sticky="w", padx=5, pady=5)
        
        ctk.CTkLabel(stats_grid, text="Último Accidente:", font=ctk.CTkFont(weight="bold")).grid(row=2, column=0, sticky="w", padx=5, pady=5)
        ctk.CTkLabel(stats_grid, text=ultimo_acc).grid(row=2, column=1, sticky="w", padx=5, pady=5)
        
        # Historial de accidentes
        historial_frame = ctk.CTkFrame(content_frame, fg_color=EstiloApp.COLOR_FRAMES)
        historial_frame.pack(fill="x", padx=20, pady=10)
        
        # Título sección historial
        titulo_historial = ctk.CTkLabel(
            historial_frame,
            text="HISTORIAL DE ACCIDENTES",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        titulo_historial.pack(pady=10)
        
        # Tabla de accidentes
        tabla_frame = ctk.CTkFrame(historial_frame, fg_color="transparent")
        tabla_frame.pack(fill="x", padx=20, pady=10)
        
        # Crear tabla
        columns = ("fecha_acc", "fecha_alta", "dias_baja", "dx", "ambito", "n_siniestro", "descripcion")
        headers = ("Fecha Accidente", "Fecha Alta", "Días Baja", "Diagnóstico", "Ámbito", "N° Siniestro", "Descripción")
        
        tree = ttk.Treeview(tabla_frame, columns=columns, show="headings")
        
        # Configurar columnas
        for i, col in enumerate(columns):
            tree.heading(col, text=headers[i])
            tree.column(col, width=100)
        
        # Ajustar anchos específicos
        tree.column("dx", width=200)
        tree.column("ambito", width=100)
        tree.column("descripcion", width=200)
        
        # Insertar datos
        for acc in result['accidentes']:
            fecha_acc = acc['fecha_acc'].strftime('%d-%m-%Y') if acc['fecha_acc'] else ''
            fecha_alta = acc['fecha_alta'].strftime('%d-%m-%Y') if acc['fecha_alta'] else ''
            
            # Truncar diagnóstico para la tabla
            dx_truncado = acc['dx'][:50] + '...' if acc['dx'] and len(acc['dx']) > 50 else acc['dx']
            
            tree.insert('', 'end', values=(
                fecha_acc,
                fecha_alta,
                acc['dias_baja'],
                dx_truncado,
                acc['ambito'],
                acc['n_siniestro'],
                acc['descripcion']
            ))
        
        # Scrollbars para la tabla
        vsb = ttk.Scrollbar(tabla_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tabla_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        
        # Botones de acción
        botones_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        botones_frame.pack(fill="x", padx=20, pady=20)
        
        # Botón para imprimir
        imprimir_btn = ctk.CTkButton(
            botones_frame,
            text="Imprimir",
            command=lambda: self._imprimir_informe(informe_window)
        )
        imprimir_btn.pack(side="left", padx=10)
        
        # Botón para exportar a PDF
        exportar_btn = ctk.CTkButton(
            botones_frame,
            text="Exportar a PDF",
            command=lambda: self._exportar_informe_pdf(result)
        )
        exportar_btn.pack(side="left", padx=10)
        
        # Botón para cerrar
        cerrar_btn = ctk.CTkButton(
            botones_frame,
            text="Cerrar",
            command=informe_window.destroy
        )
        cerrar_btn.pack(side="right", padx=10)
        
        # Centrar ventana
        informe_window.update_idletasks()
        width = informe_window.winfo_width()
        height = informe_window.winfo_height()
        x = (informe_window.winfo_screenwidth() // 2) - (width // 2)
        y = (informe_window.winfo_screenheight() // 2) - (height // 2)
        informe_window.geometry(f'{width}x{height}+{x}+{y}')


    def mostrar_descripcion_completa(self, event):
        """Mostrar la descripción completa del accidente seleccionado"""
        try:
            # Obtener el ítem seleccionado
            item = self.tree.identify_row(event.y)
            if not item:
                return
                
            # Obtener los valores del ítem
            values = self.tree.item(item, "values")
            if not values or len(values) < 5:
                return
                
            # Mostrar la descripción en un diálogo
            descripcion = values[4]
            self.mostrar_mensaje("Descripción Completa", descripcion, "info")
        except Exception as e:
            self.logger.error(f"Error al mostrar descripción completa: {str(e)}")
            self.mostrar_mensaje("Error", f"No se pudo mostrar la descripción: {str(e)}", "error")

    def limpiar_campos(self):
        """Limpiar todos los campos del formulario y el treeview"""
        try:
            # Limpiar campos del formulario
            if hasattr(self, 'entry_legajo'):
                self.entry_legajo.delete(0, 'end')
            
            if hasattr(self, 'entry_fecha_accidente'):
                try:
                    self.entry_fecha_accidente.set_date(datetime.now())
                except:
                    pass
            
            if hasattr(self, 'entry_fecha_alta'):
                try:
                    self.entry_fecha_alta.set_date(datetime.now())
                except:
                    pass
            
            if hasattr(self, 'entry_diagnostico'):
                self.entry_diagnostico.delete(0, 'end')
            
            if hasattr(self, 'entry_ambito'):
                self.entry_ambito.delete(0, 'end')
            
            if hasattr(self, 'text_descripcion'):
                self.text_descripcion.delete('1.0', 'end')
            
            if hasattr(self, 'entry_n_siniestro'):
                self.entry_n_siniestro.delete(0, 'end')
            
            # Limpiar treeview
            self._clear_treeview()
            
            # Limpiar datos de empleado
            if hasattr(self, 'nombre_empleado_label'):
                self.nombre_empleado_label.configure(text="Nombre: ")
            
            if hasattr(self, 'total_accidentes_label'):
                self.total_accidentes_label.configure(text="Total Accidentes: 0")
            
            if hasattr(self, 'ultimo_accidente_label'):
                self.ultimo_accidente_label.configure(text="⏱️ Último accidente: No registrado")
            
            if hasattr(self, 'dias_totales_label'):
                self.dias_totales_label.configure(text="📊 Total días con ART: 0 días")
            
            # Mostrar placeholder de foto
            self._mostrar_placeholder_foto()
            
            # Resetear accidente seleccionado
            self.accidente_seleccionado_id = None
            
            self.logger.info("Campos limpiados correctamente")
        except Exception as e:
            self.logger.error(f"Error al limpiar campos: {str(e)}")
            self.mostrar_mensaje("Error", f"Error al limpiar campos: {str(e)}", "error")

    def _toggle_fecha_alta(self):
        """Habilitar/deshabilitar el campo de fecha de alta según el checkbox"""
        try:
            if self.var_en_curso.get():
                # Si está en curso, deshabilitar el campo de fecha de alta
                self.entry_fecha_alta.set_date(None)
                self.entry_fecha_alta.configure(state="disabled")
            else:
                # Si no está en curso, habilitar el campo de fecha de alta
                self.entry_fecha_alta.configure(state="normal")
        except Exception as e:
            self.logger.error(f"Error al cambiar estado de fecha de alta: {str(e)}")

# Código para ejecutar la aplicación directamente
if __name__ == "__main__":
    try:
        print("\n🚀 Iniciando módulo de ART...")
        app = AplicacionART()
        app.run()
    except Exception as e:
        print(f"❌ Error crítico: {str(e)}")
        logging.basicConfig(
            filename="logs/errores_críticos_art.log",
            level=logging.CRITICAL,
            format='%(asctime)s - [%(levelname)s]: %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        logging.critical(f"Error crítico: {str(e)}\n{traceback.format_exc()}")